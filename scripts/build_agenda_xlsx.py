"""Build BCOC 2026 Conference Agenda as a color-coded Excel workbook.

Generates one master "Full Agenda" sheet plus a sheet per day, with
consistent BCOC navy/gold styling and color-coded session types.
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT = "/Users/drsmith/Documents/BCOC Website Info/docs/BCOC_2026_Conference_Agenda.xlsx"

# ---------- color palette ----------
NAVY   = "14213D"
GOLD   = "C9A84C"
GOLDLT = "FAEBC7"
RED    = "B82040"
GREEN  = "0A7E4A"
GRAY   = "BFBDB7"
TAN    = "EFE6CE"
WHITE  = "FFFFFF"
NAVYLT = "E5E8F0"
CREAM  = "FBF7EE"

TYPE_FILL = {
    "Keynote":    RED,
    "Session":    NAVY,
    "Meal":       GOLD,
    "Break":      GRAY,
    "Social":     GREEN,
    "Ceremony":   GOLD,
    "Exhibit":    TAN,
    "Sponsor":    NAVYLT,
}
TYPE_TEXT = {
    "Keynote":    WHITE,
    "Session":    WHITE,
    "Meal":       NAVY,
    "Break":      NAVY,
    "Social":     WHITE,
    "Ceremony":   NAVY,
    "Exhibit":    NAVY,
    "Sponsor":    NAVY,
}

THIN = Side(style="thin", color="D5CFBE")
BOX  = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# ---------- agenda data ----------
HEADER = ["Time", "Type", "Title", "Speaker(s) / Host", "Department / Affiliation", "Sponsor"]

DAYS = [
    {
        "tab":   "Sunday",
        "title": "Sunday, August 2 — Arrival, Setup & Community",
        "rows": [
            ("All Day",            "Exhibit", "Check-In & Registration",                     "BCOC Registration Desk", "", ""),
            ("All Day",            "Exhibit", "Vendor / Sponsor Exhibit Hall Setup",         "BCOC Exhibits Team", "", ""),
            ("2:00 – 5:00 PM",     "Social",  "Community Give-Back Volunteer Event",         "BCOC Service Committee", "Atlanta, Georgia", ""),
            ("7:00 – 8:00 PM",     "Social",  "Host City Welcome / Informal Gathering",      "BCOC Hospitality Committee", "Loews Atlanta Hotel", ""),
        ],
    },
    {
        "tab":   "Monday",
        "title": "Monday, August 3 — Vision and Legacy",
        "rows": [
            ("All Day",            "Exhibit",  "Vendor / Sponsor Exhibit Hall Setup (cont.)", "", "", ""),
            ("7:00 – 8:15 AM",     "Meal",     "Registration & Networking Breakfast",         "Hosted by BCOC", "", "REV Fire Group *Platinum / 1-800-BoardUp *Gold"),
            ("8:00 – 8:30 AM",     "Ceremony", "Opening Ceremony (Honor Guard, National Anthem, Pledge, Moment of Silence, Opening Prayer)", "Presiding: Chief Pierre Brewton, BCOC Education Coordinator. Safety Briefing: Chief Ronald Skrine, BCOC Corresponding Secretary. Opening Prayer: Chief Jason Earl Davis, BCOC Chaplain.", "BCOC Board", ""),
            ("8:30 AM",            "Ceremony", "Introduction of E-Board / Welcome Remarks",   "Dr. Hezedean Smith", "BCOC Chair", ""),
            ("8:45 AM",            "Ceremony", "Welcome Remarks — City of Atlanta",           "Office of the Mayor", "City of Atlanta", ""),
            ("8:50 AM",            "Ceremony", "Remarks — Atlanta Fire Rescue",               "Chief Rod Smith", "Atlanta Fire Rescue", ""),
            ("8:55 AM",            "Ceremony", "Reception of Dignitaries",                    "Chief T. Pierre Brewton", "BCOC Education Coordinator", ""),
            ("9:00 – 9:30 AM",     "Session",  "Leadership, Service & Community — A Senator’s Perspective", "Sen. Raphael Warnock [TBC]", "U.S. Senate", ""),
            ("9:30 – 11:00 AM",    "Keynote",  "“But Still I Lead”",                          "Dr. I. David Daniels, PhD, CSD, VPS", "Founder & CEO, ID2 Solutions", "FirstNet *Elite Premier"),
            ("11:00 AM",           "Break",    "Group Photo — All Attendees",                 "BCOC Photography Team", "", ""),
            ("12:00 – 1:00 PM",    "Meal",     "Lunch Break",                                  "", "", "Atlantic Emergency Solutions *Platinum"),
            ("1:15 – 2:15 PM",     "Keynote",  "Leadership Through Adversity: Navigating the Fire Service as Black Chief Officers", "Chief T. Pierre Brewton, Education Coordinator", "City of Spartanburg (SC) Fire Department", "Atlantic Emergency Solutions *Platinum"),
            ("2:20 – 3:00 PM",     "Session",  "NFA Update",                                  "Erik Gabliks [TBC]", "National Fire Academy", ""),
            ("3:00 – 4:00 PM",     "Session",  "Executive Insights for Chief Fire Officers: Municipal Finance & Strategic Leadership", "Donna Gayden, CFE", "BCOC Executive Coordinator", ""),
            ("7:00 – 10:00 PM",    "Social",   "Welcome Reception",                           "BCOC Hospitality Committee", "", "American Wood Council *Gold"),
        ],
    },
    {
        "tab":   "Tuesday",
        "title": "Tuesday, August 4 — Challenges & Innovation",
        "rows": [
            ("All Day",            "Exhibit",  "Vendor / Sponsor Exhibit Hall Open",          "", "", ""),
            ("8:00 – 8:30 AM",     "Meal",     "Harmony Breakfast",                            "Hosted by BCOC", "", "Sutphen *Gold / REV Fire Group *Platinum"),
            ("8:30 – 9:30 AM",     "Keynote",  "Leading Through What You Don’t Know: Executive Blind Spots in Times of Challenge", "Chief John Alston, Fire Chief (Ret.)", "Vice-Chair BCOC", "FirstNet *Elite Premier"),
            ("9:30 – 9:35 AM",     "Sponsor",  "FirstNet Sponsor Presentation",               "FirstNet, Built with AT&T", "Elite Premier Sponsor", "FirstNet *Elite Premier"),
            ("9:35 – 10:30 AM",    "Session",  "The Top 10 Moves to Use AI to Transform Fire & Emergency Services", "Chief Steven Carter", "City of Sunrise Fire Rescue", ""),
            ("10:30 AM",           "Sponsor",  "Sponsor Presentation",                         "Platinum Sponsor Representative", "", ""),
            ("10:30 – 10:45 AM",   "Break",    "Break — Visit Vendor & Sponsor Exhibit Hall", "", "", ""),
            ("10:50 – 11:30 AM",   "Session",  "The KMP Mindset: Building Leadership Teams in Uncertain Times", "Dr. Kim McNair", "Kim McNair Productions LLC", ""),
            ("11:30 – 12:15 PM",   "Session",  "Leadership with Confidence",                  "Captain Moses Jeffries, IV", "Fire Service Leadership", ""),
            ("12:15 – 1:15 PM",    "Meal",     "Luncheon",                                     "", "", "Atlantic Emergency Solutions *Platinum / MSA-Globe *Gold"),
            ("1:15 – 2:15 PM",     "Keynote",  "Mental Health Panel: Trauma, Healing, and Transformation", "Dr. I. David Daniels, PhD, CSD, VPS", "Founder & CEO, ID2 Solutions", ""),
            ("2:15 – 2:30 PM",     "Break",    "Book Signing: Dr. Daniels — “Psychosocial Hazards Are Real!” / Visit Vendor Exhibit Hall", "", "", ""),
            ("2:30 – 3:15 PM",     "Session",  "Workshop: Care Bias in Emergency Services / Leadership Strategies", "Chief Mike Scott & Chief Calvin Brown", "Grand Fire Protection District", ""),
            ("3:15 – 3:45 PM",     "Sponsor",  "CPSE Credentialing",                          "Chief Mike Higgins", "CPSE", "CPSE"),
            ("3:45 – 4:15 PM",     "Session",  "Dollar Dollar Bill, Y’all: Strategic Grant Funding", "Chief Emanuel Washington Jr.", "Orlando Fire Department", ""),
            ("4:15 – 4:30 PM",     "Break",    "Daily Wrap-up — Visit Vendor Exhibit Hall",   "", "", ""),
            ("7:00 – 10:00 PM",    "Social",   "Opening Reception Hospitality Night",         "Dr. Jeff Buchanan", "BCOC Hospitality Committee", "FirstNet *Elite Premier"),
        ],
    },
    {
        "tab":   "Wednesday",
        "title": "Wednesday, August 5 — The Future of Leadership",
        "rows": [
            ("All Day",            "Exhibit",  "Vendor / Sponsor Exhibit Hall Open",          "", "", ""),
            ("8:00 – 8:30 AM",     "Meal",     "Networking Breakfast",                         "Hosted by BCOC", "", ""),
            ("8:30 – 9:30 AM",     "Keynote",  "Global Perspectives on Executive Fire Service Leadership", "Dr. Reggie Freeman, Fire Chief (Ret.)", "Executive Director, Fire & EMS · NEOM · Founder, The Freeman Group", "The Rev Group *Platinum"),
            ("9:30 – 10:15 AM",    "Session",  "“Shaping the Standard” — Panel: Procurement Strategies (Apparatus)", "Chief Ron Skrine", "Battalion Chief, Valdosta Fire Department", "The Rev Group *Platinum"),
            ("10:15 – 10:30 AM",   "Break",    "Coffee Break — Visit Vendor Exhibit Hall",    "", "", ""),
            ("10:15 – 11:00 AM",   "Session",  "From Academy to Degree: A Workforce Model for the Modern Fire Service", "Dr. Kellie McElroy-Hooper • Chief Antonio Burden, Fire Chief (Augusta FD)", "Augusta Technical College / Augusta Fire Department", ""),
            ("11:00 – 11:45 AM",   "Session",  "From LeaderShip to Leader of ShipS™",         "Dr. Joseph Danao II, Ed.D.", "Eastern Connecticut EMS Council", ""),
            ("11:45 – 12:15 PM",   "Break",    "Networking Break — Visit Vendor Exhibit Hall","", "", ""),
            ("12:00 – 1:00 PM",    "Meal",     "Lunch",                                        "", "", "Atlantic Emergency Solutions *Platinum"),
            ("12:00 – 1:30 PM",    "Social",   "Spouse / Partners Luncheon",                  "BCOC Spouses & Partners Committee", "", ""),
            ("2:30 – 3:30 PM",     "Keynote",  "Financial “Hot Wash” (Protecting Your Legacy)", "Chief Queen Anunay & Chief Wendell Catlet", "Las Vegas Fire & Rescue", ""),
            ("3:45 – 4:30 PM",     "Session",  "From Strategy to Execution — Turning Vision into Results", "Chief Demond Simmons", "Oakland Fire Department", ""),
            ("4:30 – 5:00 PM",     "Break",    "Networking Break — Visit Vendor Exhibit Hall","", "", ""),
            ("7:00 – 10:00 PM",    "Social",   "Evening Activities / Evening Buffet",         "BCOC Social Committee", "", "Sutphen *Gold"),
        ],
    },
    {
        "tab":   "Thursday",
        "title": "Thursday, August 6 — Membership & Development",
        "rows": [
            ("All Day",            "Exhibit",  "Vendor / Sponsor Exhibit Hall Open",          "", "", ""),
            ("8:30 – 9:30 AM",     "Keynote",  "The Fire Service in 2026: Forging the Future", "Chief John Butler", "Metro Chiefs Chair", ""),
            ("9:30 – 10:00 AM",    "Session",  "Revenue Generation",                          "William Bryant", "Mobile Area Council, Scouting America", ""),
            ("10:00 – 10:45 AM",   "Session",  "Financial Access and Acumen",                 "Jere Hawkins", "Financial Services", ""),
            ("10:45 – 11:00 AM",   "Break",    "Coffee & Networking Break",                    "", "", ""),
            ("11:00 – 11:30 AM",   "Session",  "Developing an Explorer Program",              "Carlos Coronado", "Scouting America", ""),
            ("11:30 – 12:00 PM",   "Session",  "Marketing Your Vision",                       "Maisha Land", "Marketing & Communications", ""),
            ("12:00 – 1:30 PM",    "Meal",     "Sponsor Appreciation Luncheon",                "", "", "Atlantic Emergency Solutions *Platinum"),
            ("1:30 – 2:30 PM",     "Keynote",  "The Seven Pillars of Modern Leadership: Integrating Public Safety Competencies as Black Chief Officers", "Dr. Hezedean Smith, Fire Chief (Ret.)", "BCOC Chair / GESCG CEO", "The Rev Group *Platinum"),
            ("2:30 – 4:30 PM",     "Ceremony", "BCOC Board Meeting — Bylaws Update • Strategic Plan Draft • Elections Open • Extended Q&A & Closing Remarks", "BCOC Board of Directors · Dr. Hezedean Smith, Chair", "", ""),
            ("7:00 – 10:00 PM",    "Social",   "Comedy Showcase + Karaoke Night",             "Michael “Tight Mike” Randolph & Friends", "BCOC Social Committee", "Divine 9 · FirstNet *Elite Premier"),
        ],
    },
    {
        "tab":   "Friday",
        "title": "Friday, August 7 — Closing",
        "rows": [
            ("8:00 – 9:30 AM",     "Meal",     "Harmony Breakfast",                            "Hosted by BCOC", "", "Sponsor TBD"),
            ("9:30 – 11:00 AM",    "Keynote",  "Special Presentation (Closing Keynote)",      "Closing Keynote Speaker", "BCOC Closing Speaker", ""),
            ("10:00 AM – 12:00 PM","Session",  "Open Assessment Center Mentoring",            "BCOC Board", "Black Chief Officers Committee", ""),
            ("12:00 – 1:30 PM",    "Meal",     "Lunch",                                        "Hosted by BCOC", "", ""),
            ("1:30 – 4:00 PM",     "Ceremony", "Business Meeting — Election Results • Bylaws Vote", "BCOC Board of Directors", "", ""),
            ("6:30 – 8:00 PM",     "Social",   "Oratorical Contest — $10,000 Scholarship — Atlanta Civic & School Partners • Swearing In", "BCOC Scholarship Committee", "", ""),
            ("8:00 PM",            "Social",   "Post-Contest Social",                          "BCOC Social Committee", "", ""),
        ],
    },
]

# ---------- helpers ----------
def style_header(ws, ncols):
    for col in range(1, ncols + 1):
        c = ws.cell(row=1, column=col)
        c.font = Font(name="Calibri", size=11, bold=True, color=GOLD)
        c.fill = PatternFill("solid", fgColor=NAVY)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BOX
    ws.row_dimensions[1].height = 30

def style_day_title(ws, title, ncols):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    c = ws.cell(row=1, column=1, value=title)
    c.font = Font(name="Calibri", size=16, bold=True, color=GOLD)
    c.fill = PatternFill("solid", fgColor=NAVY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 36

def write_rows(ws, rows, start_row=2):
    for r_idx, row in enumerate(rows, start=start_row):
        time_, type_, title, speaker, dept, sponsor = row
        cells = [time_, type_, title, speaker, dept, sponsor]
        for c_idx, val in enumerate(cells, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = BOX
            cell.font = Font(name="Calibri", size=10)
        # Time bold
        ws.cell(row=r_idx, column=1).font = Font(name="Calibri", size=10, bold=True, color=NAVY)
        # Type pill
        type_cell = ws.cell(row=r_idx, column=2)
        type_cell.fill = PatternFill("solid", fgColor=TYPE_FILL.get(type_, NAVY))
        type_cell.font = Font(name="Calibri", size=10, bold=True, color=TYPE_TEXT.get(type_, WHITE))
        type_cell.alignment = Alignment(horizontal="center", vertical="center")
        # Row stripe by type for whole row (light tint of type color, except where strong fill applies)
        title_cell = ws.cell(row=r_idx, column=3)
        title_cell.font = Font(name="Calibri", size=10, bold=(type_ in ("Keynote","Ceremony","Meal")))
        if type_ == "Keynote":
            for c_idx in (3, 4, 5, 6):
                ws.cell(row=r_idx, column=c_idx).fill = PatternFill("solid", fgColor="FFE9EB")
        elif type_ == "Meal":
            for c_idx in (3, 4, 5, 6):
                ws.cell(row=r_idx, column=c_idx).fill = PatternFill("solid", fgColor=GOLDLT)
        elif type_ == "Ceremony":
            for c_idx in (3, 4, 5, 6):
                ws.cell(row=r_idx, column=c_idx).fill = PatternFill("solid", fgColor=CREAM)
        elif type_ == "Social":
            for c_idx in (3, 4, 5, 6):
                ws.cell(row=r_idx, column=c_idx).fill = PatternFill("solid", fgColor="E5F3EC")
        elif type_ == "Break":
            for c_idx in (3, 4, 5, 6):
                ws.cell(row=r_idx, column=c_idx).fill = PatternFill("solid", fgColor="F2F1ED")
        elif type_ == "Sponsor":
            for c_idx in (3, 4, 5, 6):
                ws.cell(row=r_idx, column=c_idx).fill = PatternFill("solid", fgColor=NAVYLT)
        elif type_ == "Exhibit":
            for c_idx in (3, 4, 5, 6):
                ws.cell(row=r_idx, column=c_idx).fill = PatternFill("solid", fgColor=TAN)
        # Sponsor column color
        if sponsor:
            ws.cell(row=r_idx, column=6).font = Font(name="Calibri", size=10, italic=True, color=NAVY, bold=True)
        # Auto-size row height a bit for wrapping
        ws.row_dimensions[r_idx].height = 36

def set_column_widths(ws):
    widths = {1: 22, 2: 14, 3: 52, 4: 36, 5: 32, 6: 36}
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w

def add_legend(ws, start_row):
    legend_row = start_row + 2
    ws.cell(row=legend_row, column=1, value="LEGEND").font = Font(name="Calibri", size=11, bold=True, color=NAVY)
    items = list(TYPE_FILL.items())
    for i, (label, color) in enumerate(items):
        c = ws.cell(row=legend_row + 1 + (i // 4), column=1 + (i % 4) * 2)
        c.value = label
        c.fill = PatternFill("solid", fgColor=color)
        c.font = Font(name="Calibri", size=10, bold=True, color=TYPE_TEXT.get(label, WHITE))
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = BOX

# ---------- build workbook ----------
wb = Workbook()

# Cover/summary
cover = wb.active
cover.title = "Cover"
cover.column_dimensions['A'].width = 110
cover.row_dimensions[2].height = 60
cover.row_dimensions[4].height = 30
cover.row_dimensions[6].height = 22

cover["A2"] = "BCOC 40TH ANNIVERSARY CONFERENCE"
cover["A2"].font = Font(name="Calibri", size=22, bold=True, color=GOLD)
cover["A2"].fill = PatternFill("solid", fgColor=NAVY)
cover["A2"].alignment = Alignment(horizontal="center", vertical="center")

cover["A4"] = "2026 Conference Program & Agenda"
cover["A4"].font = Font(name="Calibri", size=16, bold=True, color=NAVY)
cover["A4"].alignment = Alignment(horizontal="center", vertical="center")

cover["A6"] = "August 2 – 7, 2026 · Loews Atlanta Hotel · Atlanta, Georgia"
cover["A6"].font = Font(name="Calibri", size=12, color=NAVY)
cover["A6"].alignment = Alignment(horizontal="center", vertical="center")

cover["A8"] = "Use the day tabs below to navigate. The 'Full Agenda' tab lists every session across the week. Color-coded by type:"
cover["A8"].font = Font(name="Calibri", size=11, color=NAVY)
cover["A8"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

# Legend on cover
legend_row = 10
items = list(TYPE_FILL.items())
for i, (label, color) in enumerate(items):
    col = 1 + i
    c = cover.cell(row=legend_row, column=col, value=label)
    c.fill = PatternFill("solid", fgColor=color)
    c.font = Font(name="Calibri", size=11, bold=True, color=TYPE_TEXT.get(label, WHITE))
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = BOX
    cover.column_dimensions[get_column_letter(col)].width = 14
cover.row_dimensions[legend_row].height = 28

# Full Agenda master sheet
master = wb.create_sheet("Full Agenda")
master.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(HEADER) + 1)
mc = master.cell(row=1, column=1, value="BCOC 2026 — Full Conference Agenda")
mc.font = Font(name="Calibri", size=16, bold=True, color=GOLD)
mc.fill = PatternFill("solid", fgColor=NAVY)
mc.alignment = Alignment(horizontal="center", vertical="center")
master.row_dimensions[1].height = 36

# Header row at row 2 with Day prefix
header_with_day = ["Day"] + HEADER
for col, h in enumerate(header_with_day, start=1):
    c = master.cell(row=2, column=col, value=h)
    c.font = Font(name="Calibri", size=11, bold=True, color=GOLD)
    c.fill = PatternFill("solid", fgColor=NAVY)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = BOX
master.row_dimensions[2].height = 30

# Master column widths
master_widths = {1: 12, 2: 22, 3: 14, 4: 52, 5: 36, 6: 32, 7: 36}
for col, w in master_widths.items():
    master.column_dimensions[get_column_letter(col)].width = w

# Fill master rows + per-day sheets
master_row = 3
for day in DAYS:
    # Per-day sheet
    ws = wb.create_sheet(day["tab"])
    style_day_title(ws, day["title"], len(HEADER))
    # Header at row 2
    for col, h in enumerate(HEADER, start=1):
        c = ws.cell(row=2, column=col, value=h)
        c.font = Font(name="Calibri", size=11, bold=True, color=GOLD)
        c.fill = PatternFill("solid", fgColor=NAVY)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BOX
    ws.row_dimensions[2].height = 30
    write_rows(ws, day["rows"], start_row=3)
    set_column_widths(ws)
    ws.freeze_panes = "A3"

    # Append to master
    for row in day["rows"]:
        time_, type_, title, speaker, dept, sponsor = row
        cells = [day["tab"], time_, type_, title, speaker, dept, sponsor]
        for c_idx, val in enumerate(cells, start=1):
            cell = master.cell(row=master_row, column=c_idx, value=val)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = BOX
            cell.font = Font(name="Calibri", size=10)
        # Day cell
        master.cell(row=master_row, column=1).font = Font(name="Calibri", size=10, bold=True, color=NAVY)
        master.cell(row=master_row, column=1).alignment = Alignment(horizontal="center", vertical="center")
        # Time bold
        master.cell(row=master_row, column=2).font = Font(name="Calibri", size=10, bold=True, color=NAVY)
        # Type pill
        type_cell = master.cell(row=master_row, column=3)
        type_cell.fill = PatternFill("solid", fgColor=TYPE_FILL.get(type_, NAVY))
        type_cell.font = Font(name="Calibri", size=10, bold=True, color=TYPE_TEXT.get(type_, WHITE))
        type_cell.alignment = Alignment(horizontal="center", vertical="center")
        # Light tint by type for content cells
        tint_map = {
            "Keynote": "FFE9EB",
            "Meal": GOLDLT,
            "Ceremony": CREAM,
            "Social": "E5F3EC",
            "Break": "F2F1ED",
            "Sponsor": NAVYLT,
            "Exhibit": TAN,
        }
        tint = tint_map.get(type_)
        if tint:
            for c_idx in (4, 5, 6, 7):
                master.cell(row=master_row, column=c_idx).fill = PatternFill("solid", fgColor=tint)
        if type_ in ("Keynote", "Ceremony", "Meal"):
            master.cell(row=master_row, column=4).font = Font(name="Calibri", size=10, bold=True)
        if sponsor:
            master.cell(row=master_row, column=7).font = Font(name="Calibri", size=10, italic=True, color=NAVY, bold=True)
        master.row_dimensions[master_row].height = 36
        master_row += 1

master.freeze_panes = "A3"

wb.save(OUT)
print(f"wrote {OUT}")
